"""
can_spec.py — parse the "CAN Communication Plan" Excel into a decode config
============================================================================
Turns the engineering spreadsheet into a clean, machine-wise JSON config:

    {
      "version":     "v9",
      "source_file": "CAN Communication Plan_v9.xlsx",
      "parsed_at":   "2026-06-09T12:00:00+05:30",
      "function_ids":[ {"id":"0x01","name":"Start/Pause/Stop Frame","machines":"All"}, ... ],
      "machines": {
        "1": {"name":"DrawFrame",
              "identifiers":[ {"name":"Front Roller DF","addr":"0x02"}, ... ],
              "frames":[ {"frame":"Drive Check","fn":"0x03",
                          "src":"Motherboard","src_addr":"0x01",
                          "dst":"Front Roller DF","dst_addr":"0x02"}, ... ]},
        ...
      }
    }

This is the FIRST slice of "spec-driven decoding" — it produces the config that an
admin reviews/edits on the website and saves. Wiring the saved config into the live
decoder (agent.py) is a later step.

Sources inside the workbook (verified against v9):
  * Function IDs   -> sheet "All FunctionIDs"           (name | machines | id)
  * Identifiers    -> sheet "Identifiers"               (section per machine: name | 0xNN)
  * Frames         -> sheets "Extended Identifiers(...)" (Frame Type | To | From | bit cols)
                      FI/DA/SA are computed from the 29 binary bit columns, NOT from the
                      spreadsheet's own formula cells (those read back as #VALUE!).
"""
import os
import re
from datetime import datetime, timezone, timedelta

import openpyxl

IST = timezone(timedelta(hours=5, minutes=30))

# Machine numbering used everywhere else in the project (see agent.py MACHINE_NAME).
MACHINE_NAME = {1: 'DrawFrame', 2: 'BlowCard', 3: 'FlyerFrame', 4: 'RingFrame'}

# Map an "Identifiers" sheet section header -> machine id.
_IDENT_SECTION_TO_MID = {
    'draw frame': 1,
    'carding': 2,
    'carding machine': 2,
    'flyer': 3,
    'flyer frame': 3,
    'ring frame': 4,
    'ring': 4,
}

# Map a "Frame For All Machines" Machine-Group cell -> machine id.
# (The same sheet also has an abbreviation legend at the bottom whose group cells
#  are things like 'Command'/'TRPM' — those simply don't match and are skipped.)
_FRAME_GROUP_TO_MID = {
    'draw frame': 1, 'drawframe': 1,
    'carding': 2,
    'flyer frame': 3, 'flyer': 3,
    'rd': 4, 'ring frame': 4, 'ring': 4,
}
FRAME_SHEET = 'Frame For All Machines'


def _hexbyte(v):
    """Normalize a function-id / address cell into '0xNN' (2-digit), or '' if blank."""
    if v is None:
        return ''
    s = str(v).strip()
    if not s:
        return ''
    s = s.lower().replace('0x', '')
    try:
        n = int(s, 16)
    except ValueError:
        try:
            n = int(s, 10)
        except ValueError:
            return ''
    return f'0x{n:02X}'


def _clean(v):
    return '' if v is None else str(v).strip()


# ── Function IDs ───────────────────────────────────────────────────────────────
def _parse_function_ids(wb, warnings):
    out = []
    if 'All FunctionIDs' not in wb.sheetnames:
        warnings.append("sheet 'All FunctionIDs' not found")
        return out
    ws = wb['All FunctionIDs']
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        name = _clean(row[0] if len(row) > 0 else '')
        machines = _clean(row[1] if len(row) > 1 else '')
        fid = _hexbyte(row[2] if len(row) > 2 else None)
        if not name or not fid:
            continue  # blank separators / not-implemented placeholders without an id
        out.append({'id': fid, 'name': name, 'machines': machines})
    return out


# ── Identifiers (per-machine address maps) ─────────────────────────────────────
def _parse_identifiers(wb, machines, warnings):
    if 'Identifiers' not in wb.sheetnames:
        warnings.append("sheet 'Identifiers' not found")
        return
    ws = wb['Identifiers']
    cur_mid = None
    for row in ws.iter_rows(values_only=True):
        a = _clean(row[0] if len(row) > 0 else '')   # section header (col A)
        b = _clean(row[1] if len(row) > 1 else '')   # drive name      (col B)
        c = _hexbyte(row[2] if len(row) > 2 else None)  # identifier   (col C)
        if a:
            key = a.lower()
            cur_mid = _IDENT_SECTION_TO_MID.get(key)
            if cur_mid is None:
                # tolerate header variations like "Draw Frame " / "Carding Machine "
                for frag, mid in _IDENT_SECTION_TO_MID.items():
                    if frag in key:
                        cur_mid = mid
                        break
            continue
        if cur_mid and b and c:
            machines.setdefault(cur_mid, _blank_machine(cur_mid))
            machines[cur_mid]['identifiers'].append({'name': b, 'addr': c})


# ── Frames + data-byte layout (sheet "Frame For All Machines") ─────────────────
# Columns (0-based): 0 Machine Group | 1 Frame Type | 2 Source(From) | 3 Dest(To) |
#   4 DLC | 5 ACK? | 6 FI(Hex) | 7 FI(Dec) | 8 CAN Id(Hex) | 9 CAN Id(Dec) |
#   10..21 DB0..DB11 (the meaning of each data byte — "what data is sent").
_DB0_COL = 10
_DB_LAST_COL = 21   # DB11


def _addr_from_canid(can_hex):
    """Derive FI / DestAddr / SrcAddr from the 29-bit extended CAN id.

    e.g. 0x06010201 -> FI=0x01, DA=0x02, SA=0x01
    (FI = bits23-16, DA = bits15-8, SA = bits7-0).
    """
    try:
        cid = int(str(can_hex).strip().lower().replace('0x', ''), 16) & 0x1FFFFFFF
    except (ValueError, AttributeError):
        return '', '', ''
    return f'0x{(cid >> 16) & 0xFF:02X}', f'0x{(cid >> 8) & 0xFF:02X}', f'0x{cid & 0xFF:02X}'


def _data_bytes(row, dlc):
    """Collect DB0..DBn field names. Keep through DLC (or last named byte)."""
    raw = [_clean(row[c]) if c < len(row) else '' for c in range(_DB0_COL, _DB_LAST_COL + 1)]
    last = -1
    for i, v in enumerate(raw):
        if v:
            last = i
    n = dlc if isinstance(dlc, int) and dlc > 0 else last + 1
    n = max(n, last + 1)          # never truncate a named byte
    return raw[:n]


def _parse_frames(wb, machines, warnings):
    if FRAME_SHEET not in wb.sheetnames:
        warnings.append(f"sheet '{FRAME_SHEET}' not found — frames not parsed")
        return
    ws = wb[FRAME_SHEET]
    frame_type = src = dst = ''
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        group = _clean(row[0] if len(row) > 0 else '')
        mid = _FRAME_GROUP_TO_MID.get(group.lower())
        if mid is None:
            continue  # banner row (CARDING/…) or the abbreviation legend at the bottom
        # forward-fill merged cells (Frame Type / Source / Destination)
        if len(row) > 1 and _clean(row[1]):
            frame_type = _clean(row[1])
        if len(row) > 2 and _clean(row[2]):
            src = _clean(row[2])
        if len(row) > 3 and _clean(row[3]):
            dst = _clean(row[3])
        can_hex = _clean(row[8] if len(row) > 8 else '')
        fi = _hexbyte(row[6] if len(row) > 6 else None)
        if not can_hex and not fi:
            continue  # blank spacer row inside a machine section
        fi_c, da, sa = _addr_from_canid(can_hex)
        if not fi:
            fi = fi_c
        dlc = row[4] if len(row) > 4 else None
        ack = _clean(row[5] if len(row) > 5 else '')
        can_id = can_hex if can_hex.lower().startswith('0x') else (('0x' + can_hex) if can_hex else '')
        machines.setdefault(mid, _blank_machine(mid))
        machines[mid]['frames'].append({
            'frame': frame_type, 'fn': fi,
            'can_id': can_id,
            'src': src, 'src_addr': sa,
            'dst': dst, 'dst_addr': da,
            'dlc': dlc if isinstance(dlc, int) else _clean(dlc),
            'ack': ack,
            'data': _data_bytes(row, dlc),
        })


def _blank_machine(mid):
    return {'name': MACHINE_NAME.get(mid, f'M{mid}'), 'identifiers': [], 'frames': []}


def _detect_version(path):
    m = re.search(r'_v(\d+)', os.path.basename(path), re.IGNORECASE)
    return f'v{m.group(1)}' if m else ''


def parse_plan(xlsx_path):
    """Parse the CAN Communication Plan workbook into the decode-config dict.

    Never raises on a single bad sheet — collects human-readable notes in
    config['warnings'] so the admin UI can show what (if anything) went wrong.
    """
    warnings = []
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        function_ids = _parse_function_ids(wb, warnings)
        machines = {}
        _parse_identifiers(wb, machines, warnings)
        _parse_frames(wb, machines, warnings)
    finally:
        wb.close()

    # ensure all four machine slots exist and stringify keys for JSON
    machines_out = {}
    for mid in (1, 2, 3, 4):
        machines_out[str(mid)] = machines.get(mid, _blank_machine(mid))

    return {
        'version': _detect_version(xlsx_path),
        'source_file': os.path.basename(xlsx_path),
        'parsed_at': datetime.now(IST).isoformat(timespec='seconds'),
        'function_ids': function_ids,
        'machines': machines_out,
        'warnings': warnings,
    }


if __name__ == '__main__':
    import sys, json
    path = sys.argv[1] if len(sys.argv) > 1 else 'CAN Communication Plan_v9.xlsx'
    cfg = parse_plan(path)
    print(json.dumps(cfg, indent=2, ensure_ascii=False))
